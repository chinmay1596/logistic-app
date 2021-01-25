from datetime import datetime
from distlib.util import cached_property
from django.http import HttpResponse
from django.views.generic.base import View
from django.views.generic.detail import SingleObjectMixin
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from apps.commons.utils.commons import nested_getattr


class ExportBase(SingleObjectMixin, View):
    http_method_names = ['get', 'option', 'head', 'trace']
    related_names = {}

    def export_data(self):
        return self.get_queryset()

    def get_related_names(self):
        return self.related_names

    @cached_property
    def get_column(self):
        related_name = self.get_related_names()
        return related_name.keys()

    @cached_property
    def get_rows(self):
        related_name = self.get_related_names()
        return related_name.values()

    def get(self, request, *args, **kwargs):
        return self.get_workbook(self.export_data())


class ExportToExcel(ExportBase):
    """
    This mixin is used to export data into excel file.
    To request for export file one must use get method.

    Required parameters:
    1. related_names = {'display_name': 'database_field'}
    2. queryset = Default queryset

    """

    def get_response(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={name}-{date}.xlsx'.format(
            name=self.__class__.__module__.split('.')[1],
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        return response

    def get_workbook(self, queryset):
        workbook = Workbook()
        worksheet = workbook.active
        line_used = 1

        for header_index, header in enumerate(self.get_column, start=1):
            cell = worksheet.cell(row=line_used, column=header_index, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CFD8DC', end_color='CFD8DC', fill_type='solid')

        line_used += 1
        worksheet.freeze_panes = f'A{line_used}'

        _ = self.fill_data(worksheet, queryset, line_used)

        response = self.get_response()
        workbook.save(response)
        return response

    def fill_data(self, worksheet, queryset, line_used):
        for instance in queryset:
            for column_index, row in enumerate(self.get_rows, start=1):
                worksheet.cell(row=line_used, column=column_index, value=nested_getattr(instance, row))
            line_used += 1

        return line_used
