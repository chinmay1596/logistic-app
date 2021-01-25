from zipfile import BadZipFile

from django.contrib import messages
from django.core.exceptions import ValidationError
from django.http import HttpResponse, HttpResponseRedirect
from django.views.generic.base import View
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook

from apps.commons.exception.exceptions import FileImportError
from apps.commons.utils.commons import inverse_mapping


class BaseForFileImport(View):
    queryset_fields_map = None
    failed_url = None
    success_url = None
    import_form_class = None
    non_mandatory_field_value = None

    def get_import_form_class(self):
        """
        Serializer class used to import
        default to cls.serializer_class
        """
        return (
                self.import_form_class
                or self.__class__.form_class
        )

    def get_import_fields(self):
        assert hasattr(self, 'import_fields'), "You must specify import " \
                                               "field to generate sample file."
        return self.import_fields

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        import_fields = self.get_import_fields()
        # if model field is given it sets model_fields_map value else
        # it generates its default model_fields_map values from import_fields
        self.model_fields_map = getattr(
            self,
            'model_fields_map',
            None
        ) or dict(
            zip(
                import_fields,
                list(
                    map(
                        lambda fields: fields.lower().replace(' ', '_'),
                        import_fields
                    )
                )
            )
        )
        self.validate_model_fields_map(self.model_fields_map)
        self.serializer_field_map = inverse_mapping(self.model_fields_map)

    def get_failed_url(self):
        if not self.failed_url:
            raise AssertionError(
                "'{}' should either include a `failed_url` attribute, "
                "or override the `get_failed_url()` method.".format(self.__class__.__name__)
            )
        return self.failed_url

    def get_success_url(self):
        if not self.success_url:
            raise AssertionError(
                "'{}' should either include a `success_url` attribute, "
                "or override the `get_success_url()` method.".format(self.__class__.__name__)
            )
        return self.success_url

    def get_queryset_fields_map(self):
        return self.queryset_fields_map

    def get_many_to_many_fields(self):
        if hasattr(self, 'many_to_many_fields'):
            return getattr(self, 'many_to_many_fields')
        return []

    def get_non_mandatory_field_value(self):
        return self.non_mandatory_field_value

    @classmethod
    def add_validators(cls, workbook, worksheet, queryset_fields_map, serializer_field_map):
        if queryset_fields_map:
            key_qs = queryset_fields_map
            for key in key_qs:
                worksheet.add_data_validation(
                    cls.get_dv(serializer_field_map.get(key), key_qs.get(key), workbook)
                )
            return worksheet

    @classmethod
    def get_dv(cls, fieldname, qs, workbook):
        export_fields = cls.import_fields
        index = cls.get_index(fieldname, export_fields)
        slug_field = getattr(cls, 'slug_field_for_sample', 'slug')
        return cls.get_dv_for_qs(qs, index, fieldname, workbook, field=slug_field)

    @staticmethod
    def get_index(field, l):
        return l.index(field) + 1

    @staticmethod
    def get_dv_for_qs(qs, index, fieldname, workbook, field='slug'):
        fieldname = fieldname.lower().replace(' ', '_')
        if isinstance(qs, list):
            autocomplete = qs
        else:
            autocomplete = list(qs.values_list(field, flat=True))
        # v = ",".join(autocomplete)
        sheet = workbook.create_sheet(fieldname)
        max_row = len(autocomplete)
        for value in autocomplete:
            sheet.append([value])

        cl = get_column_letter(index)

        # dv = DataValidation(type="list", formula1=f'"{v}"', allow_blank=False)

        dv = DataValidation(type="list", formula1=f'{fieldname}!$A$1:$A${max_row}')
        dv.add(f'{cl}2:{cl}1048576')

        dv.error = 'Your entry is not in the list'
        dv.errorTitle = 'Invalid Entry'

        sheet.sheet_state = 'hidden'
        sheet.protection.sheet = True
        sheet.protection.password = f"{fieldname}{max_row}"
        return dv

    @staticmethod
    def validate_model_fields_map(field_map):
        keys = field_map.keys()
        values = field_map.values()
        assert len(keys) == len(set(values)), "Values must be unique for model_fields_map"


class SampleDownloadForImportMixin:
    sample_file_name = 'import'

    def get_import_fields(self):
        assert hasattr(self, 'import_fields'), "You must specify import " \
                                               "field to generate sample file."
        return self.import_fields

    def generate_sample_for_import(self):
        values = getattr(self, 'values', None)
        workbook = Workbook()
        worksheet = workbook.active
        import_fields = self.get_import_fields()
        worksheet.append(import_fields)
        worksheet.freeze_panes = 'A2'

        for i in range(1, len(import_fields) + 1):
            worksheet.column_dimensions[get_column_letter(i)].width = 30

        if values:
            worksheet.append(values)

        self.add_validators(
            workbook, worksheet,
            self.get_queryset_fields_map(), self.serializer_field_map
        )

        return save_virtual_workbook(workbook)

    def get_sample_file_name(self):
        return f'{self.sample_file_name}_{self.__class__.__module__.split(".")[1]}'

    def get(self, request, *args, **kwargs):
        response = HttpResponse(
            content=self.generate_sample_for_import(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; ' \
                                          f'filename={self.get_sample_file_name()}_sample.xlsx'
        return response


class NormalFileImportBaseMixin(BaseForFileImport):
    def start_import(self, file):
        try:
            workbook = load_workbook(file, read_only=True)
        except (OSError, BadZipFile):
            raise FileImportError('The file seems to be corrupted')

        # checking whether the file is empty
        if len(workbook.sheetnames) > 0:
            sheet_name = workbook.sheetnames[0]
        else:
            raise FileImportError('The file is empty.')

        worksheet = workbook[sheet_name]
        return list(self.extract_data_for_import(worksheet))

    def extract_data_for_import(self, worksheet):
        headers = None
        form_class = self.get_import_form_class()
        self.validate_data(worksheet)
        for index, row in enumerate(worksheet):
            row_data = [cell.value for cell in row]
            if not any(row_data):
                break
            if index == 0:
                headers = [self.model_fields_map.get(header) for header in row_data]
                self.validate_header(headers)
            else:
                data = dict(zip(headers, row_data))
                non_mandatory_field_value = self.get_non_mandatory_field_value()
                if non_mandatory_field_value:
                    none_fields = [key for key, value in non_mandatory_field_value.items()]
                    update_data = {
                        key: value for key, value in non_mandatory_field_value.items()
                        if key in none_fields
                    }
                    data.update(update_data)
                yield form_class(data=data)

    def reformat_errors(self, errors):
        """
        Convert nested errors to flat errors
        :param errors: Error list
        :return: Flat Error list
        """
        new_errors = list()
        for error in errors:
            error_dict = dict()
            for key, value in error.items():
                if isinstance(value, dict):
                    for k, v in value.items():
                        error_dict.update({self.serializer_field_map.get(k, k): v})
                else:
                    error_dict.update({self.serializer_field_map.get(key, key): value})
            new_errors.append(error_dict)
        return new_errors

    def validate_header(self, headers):
        fields = set(self.import_fields)
        if fields.issubset(set(headers)):
            raise FileImportError('Could not import holidays. ')

    @staticmethod
    def validate_data(worksheet):
        if worksheet.max_row <= 1:
            raise FileImportError("No data found for the headers.")


class NormalFileImportMixin(NormalFileImportBaseMixin, SampleDownloadForImportMixin):
    """
        This mixin is used for generating sample files for user as reference
        as well as helps to import data from file to database.

        For proper functioning of this mixin, we need to specify import_fields
        and model_fields_map as class variables. Here import_fields is mandatory
        where as model_fields_map is optional. We need to specify these fields
        within class where we extend this class.

        Example:

        :cvar import_fields:
            This field has all the fields needed to be export
            :type import_fields: list
            example:
            import_fields = [
                'NAME',
                'CATEGORY',
                'START DATE',
                'END DATE',
                'GENDER',
                'DIVISION',
                'RELIGION',
                'ETHNICITY',
                'LOWER_AGE',
                'UPPER_AGE',
                'DESCRIPTION'
            ]
        :cvar values:
            This field holds sample value for sample file. This field in non mandatory field
            :type values: list
            example:
            values = [
                'Dashain',
                'Public',
                '2019-01-01',
                '2019-01-01',
                'All',
                'IT',
                'Hinduism',
                'Tamang',
                16,
                99,
                'Holiday_description'
            ]
        :cvar model_fields_map: Non Mandatory field
            This field holds map between import_fields and model fields
            :type model_fields_map: dict
            example:
                model_fields_map = {
                    'NAME': 'name',
                    'CATEGORY': 'order_type',
                    'START DATE': 'start_date',
                    'END DATE': 'end_date',
                    'GENDER': 'gender',
                    'DIVISION': 'division',
                    'RELIGION': 'religion',
                    'ETHNICITY': 'ethnicity',
                    'LOWER_AGE': 'lower_age',
                    'UPPER_AGE': 'upper_age',
                    'DESCRIPTION': 'description',
                }

            model_fields_map is dict containing mapping between import_fields and
            model_fields. Key for model_fields_map dict represent import_fields where as
            Value for model_fields_map dict represents model_field for that module.

            Final output file as a sample.
            :```````````````````:````````````````````````:
            : NAME              : DESCRIPTION            :
            :```````````````````:````````````````````````:
            :..................:.........................:
        :cvar queryset_field_map:
            This field is used to add list for choice
            :type queryset_field_map: list
            example:
            queryset_field_map = [
                'division': 'list of choice or queryset'
            ]
        :cvar failed_url: (Mandatory Field)
            This field is used within notification to redirect user in frontend side.
            :type failed_url: str
            example:
                failed_url='/admin/org-slug/organization/settings/holiday/?status=failed'
        :cvar success_url: (Mandatory Field)
            This field is used within notification to redirect user in frontend side.
            :type failed_url: str
            example:
                failed_url='/admin/org-slug/organization/settings/holiday/?status=success'
        """

    def post(self, request, *args, **kwargs):
        file = request.FILES['import-file']
        try:
            forms = self.start_import(file)
        except FileImportError as e:
            messages.error(request, e)
            return HttpResponseRedirect(self.get_failed_url())

        if all([form.is_valid() for form in forms]):
            for form in forms:
                form.save()
            messages.success(request, 'Successfully Uploaded data.')
            return HttpResponseRedirect(self.get_success_url())

        messages.error(request, 'Unable to upload data from file.')
        return HttpResponseRedirect(self.get_failed_url())
